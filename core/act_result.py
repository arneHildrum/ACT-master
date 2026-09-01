# Copyright (c) Meta Platforms, Inc. and affiliates.
# All rights reserved.
#
# This source code is licensed under the license found in the
# LICENSE file in the root directory of this source tree.

import datetime
import getpass
import sys
from dataclasses import dataclass, field
from enum import Enum
from typing import Any

import pint
import yaml
from act.core.carbon import Carbon, SourceType
from act.core.device_data import DEVICE_FIELDS, MODEL, NAME
from act.core.utils.units import kg, units
from openpyxl import Workbook

SUMMARY_SHEET = "Summary"
CARBON_SHEET = "Carbon"
BOM_SHEET = "BOM"


@dataclass
class ACTResult:
    """Data class for storing ACT analysis results.

    Attributes:
        carbon_by_device (dict[str, Carbon]): Carbon emissions mapped by device name.
        timestamp (str): Timestamp when the analysis was run.
        cl_args (list): Command line arguments used for the analysis.
        duty_cycle (float): Device utilization rate.
        life_cycle (pint.Quantity): Expected hardware life cycle.
        total_carbon (Carbon): Total carbon emissions across all devices.
        carbon_by_category (dict[SourceType, pint.Quantity]): Carbon emissions by source type.
        metrics (dict): Computed metrics from the analysis.
        weight_unit (pint.Quantity): Unit of weight for reporting.
        bom (Any): Bill of materials used in the analysis.
    """

    carbon_by_device: dict[str, Carbon] = field(default_factory=dict)
    timestamp: str = None
    cl_args: list = field(default_factory=list)
    duty_cycle: float = None
    life_cycle: pint.Quantity = None
    total_carbon: Carbon = None
    carbon_by_category: dict[SourceType, pint.Quantity] = field(default_factory=dict)
    metrics: dict = field(default_factory=dict)
    weight_unit: pint.Quantity = kg
    bom: Any = None
    op_power: pint.Quantity = None                                                                                  # Added missing attribute for operational power
    op_ci: str = None
    op_year: int = None

    def _carbon_from_dict(self, data: dict[str, str]):
        """Convert a dictionary of string values to a Carbon object.

        Args:
            data (dict[str, str]): Dictionary mapping source type names to carbon amounts.

        Returns:
            Carbon: Carbon object initialized from the dictionary.
        """
        return Carbon(result_dict={SourceType(k): units(v) for k, v in data.items()})

    def __post_init__(self):
        # construct the Carbon values if loading from yaml
        def cast_to_carbon(results):
            for key, data in results.items():
                if not isinstance(data, Carbon):
                    results[key] = self._carbon_from_dict(data)

        cast_to_carbon(results=self.carbon_by_device)

        # populate telemetry args
        now = datetime.datetime.now()
        self.timestamp = now.strftime("%m/%d/%Y %H:%M:%S")
        self.cl_args = " ".join(sys.argv)

    def set_carbon_by_device(self, carbon_by_device):
        """Set the carbon emissions by device and recalculate totals.

        Args:
            carbon_by_device (dict[str, Carbon]): Dictionary mapping device names to carbon emissions.
        """
        self.carbon_by_device = carbon_by_device
        self.recalculate()

    def get_carbon_by_category(self):
        """Generate a report of carbon emissions by source category.

        Returns:
            dict[str, pint.Quantity]: Dictionary mapping source type names to carbon amounts.
        """
        # generate the result report by category
        result_by_cat_dict = {}
        for src in SourceType:
            result_by_cat_dict[src.name] = self.total_carbon.partial(src).to(
                self.weight_unit
            )
        return result_by_cat_dict

    def export_yaml(self, export_file):
        """Export results to a YAML file.

        Args:
            export_file (str): Path to the output YAML file.
        """
        with open(export_file, "w") as handle:
            yaml.dump(self.as_str_dict(), handle)

    def as_str_dict(self) -> dict:
        """Convert results to a dictionary with string values.

        Returns:
            dict: Dictionary representation of results with stringified values.
        """
        ret_dict = {
            "carbon_by_device": {
                dev: data.as_str_dict() for dev, data in self.carbon_by_device.items()
            },
            "timestamp": self.timestamp,
            "cl_args": self.cl_args,
            "duty_cycle": self.duty_cycle,
            "life_cycle": str(self.life_cycle),
            "total_carbon": self.total_carbon.as_str_dict(),
            "carbon_by_category": {
                cat: str(amt.to(self.weight_unit))
                for cat, amt in self.carbon_by_category.items()
            },
            "metrics": {k: str(v) for k, v in self.metrics.items()},
            "op_power": str(self.op_power),                                                                                             # Added operational power to the exported dictionary so it appears in the YAML output
            "op_ci": self.op_ci,                                                                                                        # Added operational carbon intensity to the exported dictionary so it appears in the YAML output
            "op_year": self.op_year,                                                                                                    # Added operational year to the exported dictionary so it appears in the YAML output                                               
        }

        return ret_dict

    def get_carbon_by_device(self) -> dict:
        """Get the carbon emissions by device.

        Returns:
            dict[str, Carbon]: Dictionary mapping device names to carbon emissions.
        """
        return self.carbon_by_device

    def get_fcarbon_by_device(self, weight_unit):
        """Get formatted carbon emissions by device with specified weight unit.

        Args:
            weight_unit (pint.Quantity): The weight unit to convert results to.

        Returns:
            dict[str, dict[str, str]]: Nested dictionary of device names to source types to formatted amounts.
        """
        results_by_dev_dict = {}
        for dev, carbon in self.get_carbon_by_device().items():
            dev_dict = {
                ctype.name: str(amt.to(weight_unit))
                for ctype, amt in carbon.carbon_by_type.items()
            }
            results_by_dev_dict[dev] = dev_dict
        return results_by_dev_dict

    def export_spreadsheet(self, export_file: str):
        """Export results to an Excel spreadsheet.

        Args:
            export_file (str): Path to the output Excel file.
        """
        # Create a new workbook
        wb = Workbook()

        # get rid of the initial default generated spreadsheet
        ws = wb.active
        wb.remove(ws)

        # create the report spreadsheets
        top_ws = wb.create_sheet(SUMMARY_SHEET)
        carbon_ws = wb.create_sheet(CARBON_SHEET)
        bom_ws = wb.create_sheet(BOM_SHEET)

        def write_data(ws, data):
            assert type(data) is list
            for i, row in enumerate(data, start=1):
                for j, value in enumerate(row, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = value

        def cell_data_from_results(results):
            carbon_type = [c.value for c in SourceType]
            data = [["Device", *carbon_type]]
            uniting = ["Units"] + [format(self.weight_unit.units, "~")] * len(
                carbon_type
            )
            data.append(uniting)
            for dev, carbon in results.items():
                data_line = [
                    dev,
                    *[
                        str(carbon.partial(ctype).to(self.weight_unit).m)
                        for ctype in SourceType
                    ],
                ]
                data.append(data_line)
            return data

        def cell_data_from_bom(bom):
            # order the headers so that it's name then model first
            header = [NAME, MODEL]
            for name in DEVICE_FIELDS.values():
                if name not in header:
                    header.append(name)
            assert len(header) == len(DEVICE_FIELDS), (
                f"Got header: {header}. Expected to have {DEVICE_FIELDS}"
            )

            data = [header]
            for dev in bom.devices.values():
                data_line = [getattr(dev, f) for f in header]

                # cast enums to names
                data_line = [x.value if isinstance(x, Enum) else x for x in data_line]

                # cast to string
                data_line = [str(x) for x in data_line]

                data.append(data_line)
            return data

        # export the top level telementry into the top sheet
        top_sheet_data = [
            ["ACT Analysis Results"],
            ["Author", getpass.getuser()],                                                                                    # Use for linux environment
            ["Report Generated", self.timestamp],
            ["Command Line Args", self.cl_args],
            ["Operating Power", str(self.op_power)],
            ["Operating Carbon Intensity", self.op_ci],
            ["Operating CI Year", self.op_year],
            ["Duty Cycle", self.duty_cycle],
            ["Hardware Lifetime", str(self.life_cycle)],
            ["Metrics"],
            *[[k, str(v)] for k, v in self.metrics.items()],
            ["Total Carbon"],
            *[[k, v] for k, v in self.total_carbon.as_str_dict().items()],
        ]
        write_data(ws=top_ws, data=top_sheet_data)

        # export the carbon by device data
        carbon_data = cell_data_from_results(self.carbon_by_device)
        write_data(carbon_ws, carbon_data)

        # export the bill of materials data
        bom_data = cell_data_from_bom(self.bom)
        write_data(bom_ws, bom_data)

        # Save the workbook
        wb.save(export_file)

    def recalculate(self):
        """Recalculate total carbon and carbon by category from device data."""
        self.total_carbon = sum([v for k, v in self.carbon_by_device.items()])
        self.carbon_by_category = self.get_carbon_by_category()
