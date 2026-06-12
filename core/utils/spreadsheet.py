# Copyright (c) Meta Platforms, Inc. and affiliates.
# All rights reserved.
#
# This source code is licensed under the license found in the
# LICENSE file in the root directory of this source tree.

from dataclasses import fields

from act.core.bom import BOM, DeviceData
from act.core.utils.logger import log
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

OVERVIEW_SHEET = "Overview"
BOM_SHEET = "Bill of Materials"
NAME = "name"
XLSX_EXTENSION = ".xlsx"

EXCEPTED_OVERVIEW_FIELDS = [
    "macros",
    "devices",
    "imports",
    "file",
    "cl_macros",
    "expected_carbon",
]


class Spreadsheet:
    """Utility class for importing/exporting BOM data to Excel spreadsheets.

    Attributes:
        filepath (str): Path to the spreadsheet file.
        bom_fields (list): List of BOM field names.
        device_fields (list): List of DeviceData field names.
    """

    def __init__(self, filepath: str):
        """Initialize the Spreadsheet utility.

        Args:
            filepath (str): Path to the spreadsheet file.
        """
        self.filepath = filepath

        # instantiate and get the fields
        self.bom_inst = BOM()
        self.bom_fields = [f.name for f in fields(self.bom_inst)]

        self.device_inst = DeviceData()
        self.device_fields = [f.name for f in fields(self.device_inst)]

    def export_template(self):
        """Export a template spreadsheet for users to fill in BOM data.

        Creates an Excel file with Overview and Bill of Materials sheets
        containing the appropriate field headers.
        """

        if not self.filepath.endswith(XLSX_EXTENSION):
            self.filepath = self.filepath + XLSX_EXTENSION

        # generate a template spreadsheet
        wb = Workbook()
        ws = wb.active
        wb.remove(ws)

        overview_ws = wb.create_sheet(OVERVIEW_SHEET)
        bom_ws = wb.create_sheet(BOM_SHEET)

        row_idx = 1
        for i, field in enumerate(self.bom_fields, start=1):
            if field in EXCEPTED_OVERVIEW_FIELDS:
                continue
            field_cell = overview_ws.cell(row=row_idx, column=1)
            field_cell.value = field
            row_idx += 1

        # set the column widths
        overview_ws.column_dimensions[get_column_letter(1)].width = 25
        overview_ws.column_dimensions[get_column_letter(2)].width = 50

        # generate the bill of materials template sheet
        name_cell = bom_ws.cell(row=1, column=1)
        bom_ws.column_dimensions[get_column_letter(1)].width = 40
        name_cell.value = NAME
        for i, field in enumerate(self.device_fields, start=2):
            field_cell = bom_ws.cell(row=1, column=i)
            field_cell.value = field

            bom_ws.column_dimensions[get_column_letter(i)].width = 10

        # export the template spreadsheet
        wb.save(self.filepath)

    def import_bom(self):
        """Import BOM data from a spreadsheet.

        Returns:
            BOM: A BOM object populated with data from the spreadsheet.
        """
        if not self.filepath.endswith(XLSX_EXTENSION):
            log.warning(
                f"Attempting to open spreadsheet without proper {XLSX_EXTENSION} extension. Ensure that the target file is correct."
            )

        wb = load_workbook(self.filepath)
        overview_ws = wb[OVERVIEW_SHEET]
        bom_ws = wb[BOM_SHEET]

        bom_data = dict()

        # load the bill of materials fields from the overview worksheet
        num_bom_rows = bom_ws.max_row
        for i in range(num_bom_rows):
            ws_idx = i + 1  # spreadsheet is 1 indexed
            name_cell = overview_ws.cell(row=ws_idx, column=1)
            value_cell = overview_ws.cell(row=ws_idx, column=2)

            name = name_cell.value
            value = value_cell.value

            if name not in self.bom_fields:
                log.error(
                    f"Attempted to load unknown field {name} from row {ws_idx} from {OVERVIEW_SHEET} sheet. This field will be ignored."
                )
            elif value == "" or value is None:
                bom_data[name] = None
            else:
                bom_data[name] = value

        # load the device data from the bill of materials sheet
        num_device_rows = bom_ws.max_row
        num_device_cols = bom_ws.max_column

        # extract the header fields for more efficient processing
        headers = dict()
        for i in range(0, num_device_cols):
            header_cell = bom_ws.cell(row=1, column=i + 1)
            header_value = header_cell.value
            headers[i + 1] = header_value

        devices = dict()
        for i in range(1, num_device_rows):  # first row is header fields
            row_idx = i + 1  # spreadsheet is 1 indexed
            device_dict = dict()
            name = bom_ws.cell(row=row_idx, column=1).value  # name is required

            # extract the values out of each of the column entries
            for j in range(1, num_device_cols):
                col_idx = j + 1

                cell = bom_ws.cell(row=row_idx, column=col_idx)
                field = headers[col_idx]
                device_dict[field] = cell.value

            # the name field must be specified
            if name is None or name == "":
                log.error(
                    f"Device name for bill of materials entry {row_idx} does not have a name (i.e., it's empty). A name field is required to load this device. Ignoring this entry."
                )
            else:
                devices[name] = device_dict
        # construct the bill of material data structure from loaded spreadsehet
        bom_data.update(devices=devices)
        bom = BOM(**bom_data)

        return bom
